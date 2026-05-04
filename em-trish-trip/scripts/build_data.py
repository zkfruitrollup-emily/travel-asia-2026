#!/usr/bin/env python3
"""
Build the trip JSON data files from the Em & Trish intake XLSX.

Outputs into ../data/:
  - trip.json     metadata (title, segments, dates)
  - timeline.json all events grouped by date
  - vault.json    flights, stays, on-the-ground

Run from the project root:
  python scripts/build_data.py path/to/intake.xlsx
"""
from __future__ import annotations

import datetime as dt
import json
import re
import sys
from pathlib import Path
from openpyxl import load_workbook

# ---------- helpers ----------

TZ_LABELS = {"EDT", "EST", "JST", "ICT", "PHT", "KST"}

def norm_date(v) -> str | None:
    if v is None or v == "":
        return None
    if isinstance(v, dt.datetime):
        return v.date().isoformat()
    if isinstance(v, dt.date):
        return v.isoformat()
    s = str(v).strip()
    # already ISO?
    if re.match(r"^\d{4}-\d{2}-\d{2}$", s):
        return s
    return None

def norm_time(v) -> str | None:
    """Return 'HH:MM' (24h) or a free-form label like 'TBD' / 'ALL DAY'."""
    if v is None or v == "":
        return None
    if isinstance(v, dt.time):
        return v.strftime("%H:%M")
    if isinstance(v, dt.datetime):
        return v.time().strftime("%H:%M")
    s = str(v).strip()
    # "1:30 PM" / "13:30" / "15:05 PM" tolerated
    m = re.match(r"^(\d{1,2}):(\d{2})\s*(AM|PM|am|pm)?$", s)
    if m:
        hh, mm, ampm = int(m.group(1)), int(m.group(2)), (m.group(3) or "").upper()
        if ampm == "PM" and hh < 12:
            hh += 12
        if ampm == "AM" and hh == 12:
            hh = 0
        return f"{hh:02d}:{mm:02d}"
    return s  # TBD, ALL DAY, etc.

def norm_bool(v) -> bool:
    if isinstance(v, bool):
        return v
    if v is None or v == "":
        return False
    return str(v).strip().upper() == "TRUE"

def clean(v):
    if v is None:
        return None
    if isinstance(v, str):
        s = v.strip()
        return s if s and not s.lower().startswith("todo") else None
    return v

def todo_to_null(v):
    if isinstance(v, str) and v.strip().lower().startswith("todo"):
        return None
    return v

def cell(row, idx):
    return row[idx] if idx < len(row) else None

# ---------- read sheets ----------

def read_rows(ws, header_row=4):
    headers = [c.value for c in ws[header_row]]
    rows = []
    for r in ws.iter_rows(min_row=header_row + 1, values_only=True):
        if all(c in (None, "") for c in r):
            continue
        rows.append(dict(zip(headers, r)))
    return rows

# ---------- transform ----------

def build_trip_meta():
    return {
        "title": "Em & Trish",
        "subtitle": "Bangkok & Philippines · May 7–26",
        "start_date": "2026-05-07",
        "end_date": "2026-05-26",
        "trip_begins_event_marker": "Land Bangkok (BKK)",
        "segments": [
            {"country": "USA",         "label": "New York",   "tz": "America/New_York", "tz_label": "EDT", "from": "2026-05-07", "to": "2026-05-08"},
            {"country": "Japan",       "label": "Tokyo",      "tz": "Asia/Tokyo",       "tz_label": "JST", "from": "2026-05-08", "to": "2026-05-09"},
            {"country": "Thailand",    "label": "Bangkok",    "tz": "Asia/Bangkok",     "tz_label": "ICT", "from": "2026-05-09", "to": "2026-05-13"},
            {"country": "Philippines", "label": "Philippines","tz": "Asia/Manila",      "tz_label": "PHT", "from": "2026-05-13", "to": "2026-05-26"},
        ],
    }

def build_timeline(timeline_rows):
    by_date: dict[str, dict] = {}
    seq = 0
    for r in timeline_rows:
        d = norm_date(r.get("date"))
        if not d:
            continue
        seq += 1
        ev = {
            "id":          f"{d}-{seq:03d}",
            "weekday":     clean(r.get("weekday")),
            "country":     clean(r.get("country")),
            "time":        norm_time(r.get("local_time")),
            "tz":          clean(r.get("tz")),
            "type":        clean(r.get("event_type")) or "activity",
            "title":       clean(r.get("title")),
            "location":    todo_to_null(clean(r.get("location"))),
            "status":      (clean(r.get("status")) or "confirmed").lower(),
            "locked":      norm_bool(r.get("locked")),
            "option_a":    clean(r.get("option_a")),
            "option_b":    clean(r.get("option_b")),
            "notes":       clean(r.get("notes")),
        }
        if not ev["title"]:
            continue
        day = by_date.setdefault(d, {
            "date":    d,
            "weekday": ev["weekday"],
            "country": ev["country"],
            "events":  [],
        })
        if not day["country"] and ev["country"]:
            day["country"] = ev["country"]
        day["events"].append(ev)

    # sort events within each day by time (with TBD/ALL DAY at the bottom)
    def sort_key(e):
        t = e.get("time") or "99:99"
        # treat HH:MM strictly; non-numeric goes last
        return t if re.match(r"^\d{2}:\d{2}$", t) else "99:99"

    days = []
    for d in sorted(by_date.keys()):
        day = by_date[d]
        day["events"].sort(key=sort_key)
        days.append(day)
    return {"days": days}

def build_vault(flight_rows, stay_rows, otg_rows):
    def segment_of(country):
        return "Bangkok" if country in ("Bangkok", "Thailand") else "Philippines"

    flights_by_seg: dict[str, list] = {"Bangkok": [], "Philippines": []}
    for r in flight_rows:
        seg = segment_of(clean(r.get("segment")) or "")
        f = {
            "airline":      clean(r.get("airline")),
            "code":         clean(r.get("flight_number")),
            "from":         clean(r.get("from")),
            "to":           clean(r.get("to")),
            "depart": {
                "date": norm_date(r.get("depart_date")),
                "time": norm_time(r.get("depart_local")),
                "tz":   clean(r.get("depart_tz")),
            },
            "arrive": {
                "date": norm_date(r.get("arrive_date")),
                "time": norm_time(r.get("arrive_local")),
                "tz":   clean(r.get("arrive_tz")),
            },
            "confirmation": todo_to_null(clean(r.get("confirmation_code"))),
            "source":       todo_to_null(clean(r.get("booking_source"))),
            "drive_link":   todo_to_null(clean(r.get("drive_link"))),
            "notes":        clean(r.get("notes")),
        }
        if not f["from"] or not f["to"]:
            continue
        flights_by_seg[seg].append(f)

    stays_by_seg: dict[str, list] = {"Bangkok": [], "Philippines": []}
    for r in stay_rows:
        seg = segment_of(clean(r.get("segment")) or "")
        nights = r.get("nights")
        try:
            nights = int(nights) if nights is not None else None
        except Exception:
            nights = None
        s = {
            "name":            clean(r.get("name")),
            "address":         clean(r.get("address")),
            "city":            clean(r.get("city")),
            "check_in_date":   norm_date(r.get("check_in_date")),
            "check_in_time":   norm_time(r.get("check_in_time")),
            "check_out_date":  norm_date(r.get("check_out_date")),
            "check_out_time": norm_time(r.get("check_out_time")),
            "nights":          nights,
            "confirmation":    todo_to_null(clean(r.get("confirmation_code"))),
            "source":          todo_to_null(clean(r.get("booking_source"))),
            "phone":           todo_to_null(clean(r.get("contact_phone"))),
            "drive_link":      todo_to_null(clean(r.get("drive_link"))),
            "notes":           clean(r.get("notes")),
        }
        if not s["name"]:
            continue
        stays_by_seg[seg].append(s)

    otg_by_seg: dict[str, list] = {"Bangkok": [], "Philippines": []}
    for r in otg_rows:
        country = clean(r.get("country")) or ""
        details = todo_to_null(clean(r.get("details")))
        if not details:
            continue
        seg = "Bangkok" if country == "Thailand" else ("Philippines" if country == "Philippines" else None)
        if not seg:
            continue  # skip Japan layover for v1
        otg_by_seg[seg].append({
            "country":    country,
            "category":   clean(r.get("category")),
            "title":      clean(r.get("title")),
            "details":    details,
            "drive_link": todo_to_null(clean(r.get("drive_link"))),
            "notes":      clean(r.get("notes")),
        })

    return {
        "flights":     flights_by_seg,
        "stays":       stays_by_seg,
        "on_the_ground": otg_by_seg,
    }

# ---------- main ----------

def main():
    if len(sys.argv) < 2:
        print("usage: build_data.py path/to/intake.xlsx", file=sys.stderr)
        sys.exit(1)
    src = Path(sys.argv[1])
    out_dir = Path(__file__).resolve().parent.parent / "data"
    out_dir.mkdir(parents=True, exist_ok=True)

    wb = load_workbook(src, data_only=True)
    timeline_rows = read_rows(wb["Timeline"], header_row=4)
    flight_rows   = read_rows(wb["Flights"],  header_row=4)
    stay_rows     = read_rows(wb["Stays"],    header_row=4)
    otg_rows      = read_rows(wb["On the ground"], header_row=4)

    trip = build_trip_meta()
    timeline = build_timeline(timeline_rows)
    vault = build_vault(flight_rows, stay_rows, otg_rows)

    (out_dir / "trip.json").write_text(json.dumps(trip, indent=2, ensure_ascii=False))
    (out_dir / "timeline.json").write_text(json.dumps(timeline, indent=2, ensure_ascii=False))
    (out_dir / "vault.json").write_text(json.dumps(vault, indent=2, ensure_ascii=False))

    print(f"wrote {out_dir}/trip.json")
    print(f"wrote {out_dir}/timeline.json  ({len(timeline['days'])} days)")
    print(f"wrote {out_dir}/vault.json    ({sum(len(v) for v in vault['flights'].values())} flights, "
          f"{sum(len(v) for v in vault['stays'].values())} stays, "
          f"{sum(len(v) for v in vault['on_the_ground'].values())} otg)")

if __name__ == "__main__":
    main()
